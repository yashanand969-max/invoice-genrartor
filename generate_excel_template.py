import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def create_invoice_template(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Standard_Invoice"

    # Define Colors
    NAVY_FILL = PatternFill(start_color="1A3C5E", end_color="1A3C5E", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="3A7CA5", end_color="3A7CA5", fill_type="solid")
    ACCENT_FILL = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
    
    # Define Fonts
    h1_font = Font(name="Calibri", size=18, bold=True, color="1A3C5E")
    bold_white = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    bold_black = Font(name="Calibri", size=10, bold=True, color="000000")
    std_font = Font(name="Calibri", size=10)
    
    thin = Side(border_style="thin", color="CCCCCC")
    border_std = Border(top=thin, left=thin, right=thin, bottom=thin)

    # 1. Company Header (Merged A1:I3) - Wait layout in prompt is A1:H3, we adjust to I for 9 cols
    ws.merge_cells('A1:I3')
    cell = ws['A1']
    cell.value = "YOUR COMPANY NAME\n123 Address Line, City, State 110001\nGSTIN: 22AAAAA0000A1Z5 | Ph: +91 9876543210"
    cell.font = h1_font
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 2. Invoice Meta 
    ws['A5'] = "Bill To:"
    ws['A5'].font = bold_black
    ws['E5'] = "Consignee Details:"
    ws['E5'].font = bold_black

    ws['G5'] = "Invoice No:"
    ws['G6'] = "Date:"
    ws['G7'] = "POS (State):"
    ws['G8'] = "Work Order:"
    for row in range(5, 9): ws[f'G{row}'].font = bold_black

    # 3. Table Headers (Row 11)
    headers = ["S.No", "Item Name", "Description", "SAC Code", "Brand", "Unit", "Qty", "Unit Price", "Amount"]
    for col_num, header in enumerate(headers, 1):
        c = ws.cell(row=11, column=col_num)
        c.value = header
        c.font = bold_white
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center" if header in ("S.No", "Unit", "Qty") else "left")
        c.border = border_std

    # 4. Line Items & Formulas (Rows 12-30)
    for row in range(12, 31):
        for col_num in range(1, 10):
            c = ws.cell(row=row, column=col_num)
            c.border = border_std
            if col_num in (7, 8, 9): c.alignment = Alignment(horizontal="right")
        
        # Format Price and Amount as currency
        ws[f'H{row}'].number_format = '#,##0.00'
        ws[f'I{row}'].number_format = '#,##0.00'
        
        # Formula: Amount = Qty * Price
        ws[f'I{row}'].value = f'=IF(OR(G{row}="",H{row}=""), "", G{row}*H{row})'

    # 5. Data Validation for Unit and SAC
    dv_unit = DataValidation(type="list", formula1='"Rft,Nos,Sqft,LS,Set,Mtr,Box,Lot"', allow_blank=True)
    ws.add_data_validation(dv_unit)
    dv_unit.add(f'F12:F30')

    # 6. Totals Section (Rows 32-37)
    labels = ["Electrical Subtotal:", "Outstation Charges (5%):", "Taxable Amount:", "IGST (18%):", "CGST+SGST (9%+9%):", "GRAND TOTAL:"]
    formulas = [
        "=SUM(I12:I30)",
        "=I32*0.05",
        "=I32+I33",
        "=I34*0.18",
        '="-"',
        "=SUM(I34, I35)"
    ]
    
    row_idx = 32
    for label, form in zip(labels, formulas):
        ws[f'H{row_idx}'] = label
        ws[f'H{row_idx}'].font = bold_black
        ws[f'H{row_idx}'].alignment = Alignment(horizontal="right")
        
        ws[f'I{row_idx}'] = form
        ws[f'I{row_idx}'].number_format = '#,##0.00'
        if "GRAND TOTAL" in label:
            ws[f'I{row_idx}'].font = bold_black
            ws[f'H{row_idx}'].fill = ACCENT_FILL
            ws[f'I{row_idx}'].fill = ACCENT_FILL
        row_idx += 1

    # 7. Amount in Words
    ws['A34'] = "Amount in Words:"
    ws['A34'].font = bold_black
    ws['C34'] = "Please input words manually or use SpellNumber add-in"
    ws['C34'].font = Font(italic=True, color="555555")

    # 8. Bank Details
    ws['A39'] = "Bank Details"
    ws['A39'].font = bold_black
    ws['A40'] = "Bank Name: HDFC Bank Ltd.\nAccount No: XXXXXXXXXXXXXXX\nIFSC Code: HDFC0000123"
    ws['A40'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[40].height = 45

    # 9. Terms and sig
    ws['A44'] = "Terms & Conditions"
    ws['A44'].font = bold_black
    ws['A45'] = "1. Payment due within 30 days.\n2. Goods once sold will not be taken back.\n3. Subject to jurisdiction."
    ws['A45'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[45].height = 45

    ws['H45'] = "Authorized Signatory"
    ws['H45'].font = bold_black
    ws['H45'].alignment = Alignment(horizontal="center")

    # Column Widths
    widths = {'A':5, 'B':25, 'C':30, 'D':10, 'E':15, 'F':8, 'G':8, 'H':12, 'I':15}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(filename)
    print(f"Generated {filename} successfully.")

if __name__ == "__main__":
    create_invoice_template("c:/Users/yasha/Desktop/i4u antigravity/Standardized_Invoice_Template.xlsx")
